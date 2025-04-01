import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# Function to extract all tables from the first two pages and convert to Excel
def pdf_to_excel(pdf_path, excel_path):
    with pdfplumber.open(pdf_path) as pdf:
        writer = pd.ExcelWriter(excel_path, engine='openpyxl')
        workbook = writer.book
        
        for page_num in range(3):  # Only process first 2 pages
            if page_num >= len(pdf.pages):
                break  # Stop if PDF has fewer than 2 pages
            
            page = pdf.pages[page_num]
            tables = page.extract_tables()
            
            if tables:
                for i, table in enumerate(tables):  # Extract all tables from the page
                    df = pd.DataFrame(table)
                    sheet_name = f"Page{page_num+1}_Table{i+1}"
                    df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=2)
                    
                    # Extract table title
                    text = page.extract_text()
                    if text:
                        lines = text.split('\n')
                        title = lines[i] if i < len(lines) else f"Table {i+1}"
                    else:
                        title = f"Table {i+1}"
                    
                    # Auto-adjust column width and set title
                    worksheet = writer.sheets[sheet_name]
                    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
                    worksheet.cell(row=1, column=1, value=title).alignment = Alignment(horizontal='center', vertical='center')
                    
                    for col_num, col_cells in enumerate(df.columns, start=1):
                        max_length = max(df[col_cells].astype(str).map(len).max(), len(str(col_cells))) + 2
                        worksheet.column_dimensions[get_column_letter(col_num)].width = max_length
                        
        writer.close()
        print(f"PDF data has been saved to {excel_path}")

# Example usage
pdf_to_excel("sample-tables.pdf", "excel2.xlsx")
